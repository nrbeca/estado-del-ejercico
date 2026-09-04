"""
Constructor de Tabla Dinámica Presupuestal — MAP / SICOP
==========================================================
App Streamlit que integra los reportes crudos de MAP y SICOP, deja armar
cualquier reporte tipo tabla dinámica (agregar/quitar filas, columnas y
valores, filtrar por cualquier campo de la base — incluida Unidad
Responsable, Partida y Programa por nombre) y descarga el resultado en
Excel con el formato institucional del "Estado del Ejercicio".

Cómo correrla:
    pip install -r requirements.txt
    streamlit run pivot_presupuestal_app.py

Qué hace por ti automáticamente al cargar un archivo crudo:
    - Detecta la codificación (MAP = utf-8, SICOP = latin-1) sola.
    - Construye la Partida completa en SICOP (Capítulo+Concepto+Genérica+
      Específica) igual que en tus reportes actuales.
    - Junta los catálogos de catalogs/unidades.csv, catalogs/partidas.csv
      y catalogs/programas.csv para mostrar nombres, no solo códigos, en
      los filtros y en las filas del reporte. Si un código no está en el
      catálogo, muestra el código tal cual — puedes ir agregando filas a
      esos CSV para completar la cobertura.
    - Calcula, para cada familia de importes (Original, Modificado,
      Comprometido, Ejercido, Reservas, etc.), el total "Anual" y el
      "Al periodo" (acumulado de enero al mes que elijas), igual que en
      el Estado del Ejercicio.
    - Calcula el Importe Disponible como Modificado − Ejercido − Comprometido
      (fórmula verificada contra tu archivo de ejemplo, cuadra al centavo).
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paleta y estilos — tomados directamente de formato_estado_del_ejercicio.xlsx
# ---------------------------------------------------------------------------
BURDEOS = "621333"     # encabezados de columnas
VERDE = "1E5B4F"       # encabezados de grupo ("Anual" / "Al periodo")
GRIS_TOTAL = "D9D9D9"  # fila de Total general
FORMATO_MONEDA = '_-"$"* #,##0.00_-;\\-"$"* #,##0.00_-;_-"$"* "-"??_-;_-@_-'

THIN_GRIS = Side(style="thin", color="808080")
BORDE = Border(left=THIN_GRIS, right=THIN_GRIS, top=THIN_GRIS, bottom=THIN_GRIS)
THIN_BLANCO = Side(style="thin", color="FFFFFF")
BORDE_ENCABEZADO = Border(left=THIN_BLANCO, right=THIN_BLANCO, top=THIN_BLANCO, bottom=THIN_BLANCO)

MESES = ["ENE", "FEB", "MAR", "ABR", "MAY", "JUN", "JUL", "AGO", "SEP", "OCT", "NOV", "DIC"]
NOMBRES_MES = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
               "agosto", "septiembre", "octubre", "noviembre", "diciembre"]

CATALOGS_DIR = Path(__file__).parent / "catalogs"

# ---------------------------------------------------------------------------
# Definición de familias de importes por fuente
# (columna total ya existente en el crudo, o None si hay que sumarla)
# ---------------------------------------------------------------------------
def _cols_sicop(prefijo_2letras: list[str]) -> list[str]:
    return prefijo_2letras


FAMILIAS_SICOP = {
    "ORIGINAL": (["OREN", "ORFE", "ORMR", "ORAB", "ORMY", "ORJN", "ORJL", "ORAG", "ORSE", "OROC", "ORNO", "ORDI"], "ORIGINAL"),
    "AMPLIACION": (["AMEN", "AMFE", "AMMR", "AMAB", "AMMY", "AMJN", "AMJL", "AMAG", "AMSE", "AMOC", "AMNO", "AMDI"], "AMPLIACION"),
    "REDUCCION": (["REEN", "REFE", "REMR", "REAB", "REMY", "REJN", "REJL", "REAG", "RESE", "REOC", "RENO", "REDI"], "REDUCCION"),
    "RESERVAS": (["RESERVA_ENE", "RESERVA_FEB", "RESERVA_MZO", "RESERVA_ABR", "RESERVA_MAY", "RESERVA_JUN",
                  "RESERVA_JUL", "RESERVA_AGO", "RESERVA_SEP", "RESERVA_OCT", "RESERVA_NOV", "RESERVA_DIC"], "RESERVAS"),
    "MODIFICADO_AUTORIZADO": (["MOEN", "MOFE", "MOMR", "MOAB", "MOMY", "MOJN", "MOJL", "MOAG", "MOSE", "MOOC", "MONO", "MODI"], "MODIFICADO_AUTORIZADO"),
    "COMPROMETIDO": (["COEN", "COFE", "COMR", "COAB", "COMY", "COJN", "COJL", "COAG", "COSE", "COOC", "CONO", "CODI"], "COMPROMETIDO"),
    "EJERCIDO": (["EJEN", "EJFE", "EJMR", "EJAB", "EJMY", "EJJN", "EJJL", "EJAG", "EJSE", "EJOC", "EJNO", "EJDI"], "EJERCIDO"),
    "DEVENGADO": (["DVEN", "DVFE", "DVMR", "DVAB", "DVMY", "DVJN", "DVJL", "DVAG", "DVSE", "DVOC", "DVNO", "DVDI"], "DEVENGADO"),
    "EJERCIDO_TRAMITE": (["EJTREN", "EJTRFE", "EJTRMR", "EJTRAB", "EJTRMY", "EJTRJN", "EJTRJL", "EJTRAG",
                          "EJTRSE", "EJTROC", "EJTRNO", "EJTRDI"], "EJERCIDO_TRAMITE"),
}
CLAVES_SICOP = {"modificado": "MODIFICADO_AUTORIZADO", "ejercido": "EJERCIDO", "comprometido": "COMPROMETIDO"}

FAMILIAS_MAP = {
    f"{pref}": ([f"{pref}_{m}" for m in MESES], None)
    for pref in ["ORI", "AMP", "RED", "MOD", "CONG", "DESCONG", "EJE"]
}
CLAVES_MAP = {"modificado": "MOD", "ejercido": "EJE", "comprometido": None}

NOMBRES_FAMILIA = {
    "ORIGINAL": "Original", "ORI": "Original",
    "AMPLIACION": "Ampliación", "AMP": "Ampliación",
    "REDUCCION": "Reducción", "RED": "Reducción",
    "RESERVAS": "Reservas",
    "MODIFICADO_AUTORIZADO": "Modificado", "MOD": "Modificado",
    "COMPROMETIDO": "Comprometido",
    "EJERCIDO": "Ejercido", "EJE": "Ejercido",
    "DEVENGADO": "Devengado",
    "EJERCIDO_TRAMITE": "Ejercido en trámite",
    "CONG": "Congelado", "DESCONG": "Descongelado",
}

FUENTES = {
    "MAP": "Módulo de Adecuaciones Presupuestarias (MAP)",
    "SICOP": "Sistema de Contabilidad y Presupuesto (SICOP)",
}

# Normalización de Unidad Responsable — tomada tal cual de MAPEO_UR_2026_BASE y
# FUSION_URS_2026 en nrbeca/nuevo/config.py, para que claves legadas o alternas
# se agrupen bajo el código vigente antes de buscar el nombre.
MAPEO_UR_BASE = {
    "G00": "811", "108": "810", "113": "250", "121": "260", "122": "261", "123": "262",
    "124": "263", "125": "264", "126": "265", "127": "266", "128": "267", "129": "268",
    "130": "269", "131": "270", "132": "271", "133": "272", "134": "273", "135": "274",
    "136": "275", "137": "276", "138": "277", "139": "278", "140": "279", "141": "280",
    "142": "281", "143": "282", "144": "283", "145": "284", "146": "285", "147": "286",
    "148": "287", "149": "288", "150": "289", "151": "290", "152": "291", "153": "292",
    "215": "220", "300": "225", "310": "226", "700": "227", "600": "230", "612": "231",
    "312": "232", "315": "233", "400": "235", "311": "237", "314": "245",
}
FUSION_URS = {
    "810": "119", "812": "119", "800": "120", "811": "120", "235": "250", "236": "253",
    "237": "253", "225": "900", "245": "910", "241": "911", "246": "912", "247": "912",
    "230": "920", "226": "921", "227": "922", "231": "923", "232": "924",
}


def normalizar_ur(codigo) -> str:
    """Aplica el mismo encadenamiento MAPEO_UR_2026_BASE -> FUSION_URS_2026 que
    usa tu Dashboard de Austeridad, para que un código legado o alterno caiga
    bajo el mismo código vigente que usa el catálogo de nombres."""
    clave = str(codigo).strip()
    clave = MAPEO_UR_BASE.get(clave, clave)
    clave = FUSION_URS.get(clave, clave)
    return clave


# ---------------------------------------------------------------------------
# Catálogos (código -> nombre) — se cargan de /catalogs, editables por ti
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner=False)
def cargar_catalogo(nombre_archivo: str, col_codigo: str, col_nombre: str) -> dict[str, str]:
    ruta = CATALOGS_DIR / nombre_archivo
    if not ruta.exists():
        return {}
    try:
        cat = pd.read_csv(ruta, dtype=str)
        cat[col_codigo] = cat[col_codigo].str.strip()
        return dict(zip(cat[col_codigo], cat[col_nombre]))
    except Exception:
        return {}


def etiqueta_con_nombre(codigo, catalogo: dict[str, str]) -> str:
    if pd.isna(codigo):
        return "(sin dato)"
    clave = str(codigo).strip()
    clave_num = clave.split(".")[0] if clave.replace(".", "", 1).isdigit() else clave
    nombre = catalogo.get(clave) or catalogo.get(clave_num)
    return f"{clave} — {nombre}" if nombre else clave


# ---------------------------------------------------------------------------
# Carga y preparación de datos crudos
# ---------------------------------------------------------------------------
@st.cache_data(show_spinner="Leyendo archivo...")
def cargar_crudo(archivo_bytes: bytes, nombre_archivo: str, fuente: str) -> pd.DataFrame:
    buffer = io.BytesIO(archivo_bytes)
    if nombre_archivo.lower().endswith((".xlsx", ".xls")):
        df = pd.read_excel(buffer)
    else:
        encodings = ["utf-8", "latin-1", "cp1252"]
        df = None
        for enc in encodings:
            try:
                buffer.seek(0)
                df = pd.read_csv(buffer, encoding=enc, low_memory=False)
                break
            except UnicodeDecodeError:
                continue
        if df is None:
            raise ValueError("No se pudo leer el archivo con ninguna codificación común (utf-8/latin-1/cp1252).")
    df.columns = [str(c).strip() for c in df.columns]

    if fuente == "SICOP":
        for c in ["CAPITULO", "CONCEPTO", "PARTIDA_GENERICA", "PARTIDA_ESPECIFICA"]:
            if c in df.columns:
                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
        if {"CAPITULO", "CONCEPTO", "PARTIDA_GENERICA", "PARTIDA_ESPECIFICA"}.issubset(df.columns):
            df["Partida"] = (df["CAPITULO"] * 10000 + df["CONCEPTO"] * 1000
                              + df["PARTIDA_GENERICA"] * 100 + df["PARTIDA_ESPECIFICA"])
        if "ID_UNIDAD" in df.columns:
            df["Unidad Responsable (código en la base)"] = df["ID_UNIDAD"].astype(str).str.strip()
            df["Unidad Responsable"] = df["Unidad Responsable (código en la base)"].apply(normalizar_ur)
        if "PROGRAMA_PRESUPUESTARIO" in df.columns:
            df["Programa"] = df["PROGRAMA_PRESUPUESTARIO"].astype(str).str.strip()
    elif fuente == "MAP":
        if "PARTIDA" in df.columns:
            df["Partida"] = pd.to_numeric(df["PARTIDA"], errors="coerce")
        if "UNIDAD" in df.columns:
            df["Unidad Responsable (código en la base)"] = df["UNIDAD"].astype(str).str.strip()
            df["Unidad Responsable"] = df["Unidad Responsable (código en la base)"].apply(normalizar_ur)
        if "PROGRAMA" in df.columns:
            df["Programa"] = df["PROGRAMA"].astype(str).str.strip()

    if "Partida" in df.columns:
        df["Capitulo"] = (pd.to_numeric(df["Partida"], errors="coerce") // 10000).astype("Int64")

    return df


def solo_nombre(codigo, catalogo: dict[str, str]) -> str:
    if pd.isna(codigo):
        return ""
    clave = str(codigo).strip()
    clave_num = clave.split(".")[0] if clave.replace(".", "", 1).isdigit() else clave
    return catalogo.get(clave) or catalogo.get(clave_num) or ""


def enriquecer_con_catalogos(df: pd.DataFrame) -> pd.DataFrame:
    cat_ur = cargar_catalogo("unidades.csv", "codigo_ur", "nombre_ur")
    cat_partidas = cargar_catalogo("partidas.csv", "partida", "nombre_partida")
    cat_programas = cargar_catalogo("programas.csv", "programa", "nombre_programa")
    cat_capitulos = cargar_catalogo("capitulos.csv", "capitulo", "nombre_capitulo")

    nuevas = {}
    if "Unidad Responsable" in df.columns:
        nuevas["Nombre de la Unidad Responsable"] = df["Unidad Responsable"].apply(lambda x: solo_nombre(x, cat_ur))
        nuevas["Unidad Responsable (nombre)"] = df["Unidad Responsable"].apply(lambda x: etiqueta_con_nombre(x, cat_ur))
    if "Partida" in df.columns:
        nuevas["Nombre Partida"] = df["Partida"].apply(lambda x: solo_nombre(x, cat_partidas))
        nuevas["Partida (nombre)"] = df["Partida"].apply(lambda x: etiqueta_con_nombre(x, cat_partidas))
    if "Programa" in df.columns:
        nuevas["Nombre Programa"] = df["Programa"].apply(lambda x: solo_nombre(x, cat_programas))
        nuevas["Programa (nombre)"] = df["Programa"].apply(lambda x: etiqueta_con_nombre(x, cat_programas))
    if "Capitulo" in df.columns:
        nuevas["Nombre Capítulo"] = df["Capitulo"].apply(lambda x: solo_nombre(x, cat_capitulos))
        nuevas["Capítulo (nombre)"] = df["Capitulo"].apply(lambda x: etiqueta_con_nombre(x, cat_capitulos))
    return pd.concat([df, pd.DataFrame(nuevas, index=df.index)], axis=1)


def agregar_periodos_y_disponible(df: pd.DataFrame, fuente: str, mes_corte_idx: int) -> tuple[pd.DataFrame, list[str]]:
    """Agrega columnas '<Familia> (Anual)' y '<Familia> (Al <mes>)' por cada
    familia de importes disponible, más 'Disponible (Anual)'/'Disponible (Al <mes>)'.
    Regresa el df enriquecido y la lista de columnas de valor calculadas (en orden lógico)."""
    familias = FAMILIAS_SICOP if fuente == "SICOP" else FAMILIAS_MAP
    claves = CLAVES_SICOP if fuente == "SICOP" else CLAVES_MAP
    mes_label = NOMBRES_MES[mes_corte_idx].capitalize()

    df = df.copy()
    columnas_valor = []
    for fam, (cols_mensuales, col_total_existente) in familias.items():
        presentes = [c for c in cols_mensuales if c in df.columns]
        if not presentes:
            continue
        nombre_bonito = NOMBRES_FAMILIA.get(fam, fam)
        col_anual = f"{nombre_bonito} (Anual)"
        if col_total_existente and col_total_existente in df.columns:
            df[col_anual] = pd.to_numeric(df[col_total_existente], errors="coerce").fillna(0)
        else:
            df[col_anual] = df[presentes].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1)

        cols_al_periodo = [c for c in cols_mensuales[: mes_corte_idx + 1] if c in df.columns]
        col_periodo = f"{nombre_bonito} (Al {mes_label})"
        df[col_periodo] = df[cols_al_periodo].apply(pd.to_numeric, errors="coerce").fillna(0).sum(axis=1) if cols_al_periodo else 0.0

        columnas_valor += [col_anual, col_periodo]

    # Disponible = Modificado - Ejercido - Comprometido (verificado contra el
    # archivo de ejemplo: cuadra al centavo con las columnas "Importe Disponible").
    nombre_mod = NOMBRES_FAMILIA.get(claves["modificado"])
    nombre_eje = NOMBRES_FAMILIA.get(claves["ejercido"])
    nombre_com = NOMBRES_FAMILIA.get(claves["comprometido"]) if claves["comprometido"] else None

    col_mod_anual, col_mod_periodo = f"{nombre_mod} (Anual)", f"{nombre_mod} (Al {mes_label})"
    col_eje_anual = f"{nombre_eje} (Anual)"
    if col_mod_anual in df.columns and col_eje_anual in df.columns:
        com_anual = df[f"{nombre_com} (Anual)"] if nombre_com and f"{nombre_com} (Anual)" in df.columns else 0
        com_periodo = df[f"{nombre_com} (Al {mes_label})"] if nombre_com and f"{nombre_com} (Al {mes_label})" in df.columns else 0
        df["Disponible (Anual)"] = df[col_mod_anual] - df[col_eje_anual] - com_anual
        df["Disponible (Al " + mes_label + ")"] = df[col_mod_periodo] - df[col_eje_anual] - com_periodo
        columnas_valor += ["Disponible (Anual)", f"Disponible (Al {mes_label})"]

    # "Ejercido real" (solo SICOP) = Ejercido + Devengado + Ejercido en trámite,
    # tal como lo define tu Dashboard de Presupuesto (EJERCIDO_REAL en
    # sicop_processor.py). Se agrega aparte, sin tocar "Ejercido", porque el
    # formato estándar de Estado del Ejercicio OREF usa Ejercido solo.
    if fuente == "SICOP" and all(f"{n} (Anual)" in df.columns for n in ["Ejercido", "Devengado", "Ejercido en trámite"]):
        df["Ejercido real (Anual)"] = df["Ejercido (Anual)"] + df["Devengado (Anual)"] + df["Ejercido en trámite (Anual)"]
        df[f"Ejercido real (Al {mes_label})"] = (df["Ejercido (Anual)"] + df[f"Devengado (Al periodo {mes_label})"]
                                                  + df[f"Ejercido en trámite (Al periodo {mes_label})"])
        columnas_valor += ["Ejercido real (Anual)", f"Ejercido real (Al periodo {mes_label})"]

    return df, columnas_valor


def construir_reporte_plantilla(df: pd.DataFrame, fuente: str, mes_corte_idx: int):
    """Reproduce exactamente el formato de formato_estado_del_ejercicio.xlsx:
    Unidad Responsable, Nombre UR, Partida, Nombre Partida + Autorizado/
    Reservado/Modificado/Comprometido (Anual y Al periodo) + Ejercido +
    Disponible (Anual y Al periodo). Si la base no trae alguno de estos
    conceptos (ej. MAP no tiene Reservado ni Comprometido), esa columna
    simplemente se omite. Regresa (tabla, encabezados, grupos, filas)."""
    mes_label = NOMBRES_MES[mes_corte_idx].capitalize()
    filas = [c for c in ["Unidad Responsable", "Nombre de la Unidad Responsable", "Partida", "Nombre Partida"] if c in df.columns]

    pares = [("Original", "Importe Autorizado"), ("Reservas", "Importe Reservado"),
             ("Modificado", "Importe Modificado"), ("Comprometido", "Importe Comprometido")]

    especificacion = []  # (columna_interna, encabezado, grupo)
    for base, etiqueta in pares:
        col = f"{base} (Anual)"
        if col in df.columns:
            especificacion.append((col, etiqueta, "Anual"))
    for base, etiqueta in pares:
        col = f"{base} (Al {mes_label})"
        if col in df.columns:
            especificacion.append((col, etiqueta, "Al periodo"))
    if "Ejercido (Anual)" in df.columns:
        especificacion.append(("Ejercido (Anual)", "Importe Ejercido", None))
    if "Disponible (Anual)" in df.columns:
        especificacion.append(("Disponible (Anual)", "Importe Disponible", "Anual"))
    if f"Disponible (Al {mes_label})" in df.columns:
        especificacion.append((f"Disponible (Al {mes_label})", "Importe Disponible", "Al periodo"))

    cols_internas = [c for c, _, _ in especificacion]
    if not filas or not cols_internas:
        return pd.DataFrame(), [], [], filas

    agregado = df.groupby(filas, as_index=False)[cols_internas].sum()
    fila_total = {c: "" for c in agregado.columns}
    fila_total[filas[0]] = "Total general"
    for c in cols_internas:
        fila_total[c] = agregado[c].sum()
    agregado = pd.concat([pd.DataFrame([fila_total]), agregado], ignore_index=True)
    agregado = agregado[filas + cols_internas]

    encabezados = filas + [etiqueta for _, etiqueta, _ in especificacion]

    grupos = []
    n_filas = len(filas)
    i = 0
    while i < len(especificacion):
        grupo = especificacion[i][2]
        if grupo in ("Anual", "Al periodo"):
            j = i
            while j < len(especificacion) and especificacion[j][2] == grupo:
                j += 1
            texto = "Anual" if grupo == "Anual" else "Al periodo"
            grupos.append((n_filas + i + 1, n_filas + j, texto))
            i = j
        else:
            i += 1

    return agregado, encabezados, grupos, filas


def aplicar_depuracion_sicop(df: pd.DataFrame) -> pd.DataFrame:
    """Reglas confirmadas en sicop_processor.py (nrbeca/nuevo) para el
    Estado del Ejercicio: excluir capítulo 1000 (servicios personales),
    la partida 39801, y CONTROL_OPERATIVO entre 60 y 69. Revisa que
    correspondan al reporte que quieres armar antes de activarlas."""
    dff = df.copy()
    col_cap = "Capitulo" if "Capitulo" in dff.columns else ("CAPITULO" if "CAPITULO" in dff.columns else None)
    if col_cap:
        dff = dff[dff[col_cap] != 1]
    if "Partida" in dff.columns:
        dff = dff[dff["Partida"] != 39801]
    if "CONTROL_OPERATIVO" in dff.columns:
        co = pd.to_numeric(dff["CONTROL_OPERATIVO"], errors="coerce")
        dff = dff[~co.between(60, 69)]
    return dff


def columnas_categoricas(df: pd.DataFrame, columnas_valor: list[str]) -> list[str]:
    return [c for c in df.columns if c not in columnas_valor and not pd.api.types.is_float_dtype(df[c])] or \
        [c for c in df.columns if c not in columnas_valor]


# ---------------------------------------------------------------------------
# Construcción de la tabla dinámica
# ---------------------------------------------------------------------------
def construir_pivote(df, filas, columnas, valores, filtros) -> pd.DataFrame:
    dff = df.copy()
    for col, seleccion in filtros.items():
        if seleccion:
            dff = dff[dff[col].isin(seleccion)]

    if not filas or not valores:
        return pd.DataFrame()

    if columnas:
        pivote = pd.pivot_table(dff, index=filas, columns=columnas, values=valores, aggfunc="sum", fill_value=0)
        pivote.columns = [" | ".join(str(x) for x in c) if isinstance(c, tuple) else str(c) for c in pivote.columns]
        pivote = pivote.reset_index()
    else:
        pivote = dff.groupby(filas, as_index=False)[valores].sum()

    fila_total = {c: "" for c in pivote.columns}
    if filas:
        fila_total[filas[0]] = "Total general"
    for c in pivote.columns:
        if c not in filas and pd.api.types.is_numeric_dtype(pivote[c]):
            fila_total[c] = pivote[c].sum()
    pivote = pd.concat([pd.DataFrame([fila_total]), pivote], ignore_index=True)
    return pivote


# ---------------------------------------------------------------------------
# Exportación a Excel con el formato del Estado del Ejercicio
# ---------------------------------------------------------------------------
def exportar_excel_oref(pivote: pd.DataFrame, fuente: str, linea1: str, linea2: str,
                         titulo: str, subtitulo: str, filas: list[str],
                         encabezados: list[str] | None = None,
                         grupos: list[tuple[int, int, str]] | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"
    ws.sheet_view.showGridLines = False

    n_cols = max(len(pivote.columns), 1)
    ultima_col = get_column_letter(n_cols)
    encabezados = encabezados or [str(c) for c in pivote.columns]

    c = ws.cell(row=1, column=n_cols, value=linea1)
    c.font = Font(name="Arial", size=12, bold=True)
    c.alignment = Alignment(horizontal="right")

    c = ws.cell(row=2, column=n_cols, value=linea2)
    c.font = Font(name="Arial", size=11, bold=True)
    c.alignment = Alignment(horizontal="right")

    ws.merge_cells(f"A5:{ultima_col}5")
    c = ws["A5"]; c.value = titulo
    c.font = Font(name="Calibri", size=14, bold=True)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[5].height = 18.75

    ws.merge_cells(f"A6:{ultima_col}6")
    c = ws["A6"]; c.value = subtitulo
    c.font = Font(name="Calibri", size=12, italic=True)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[6].height = 15.75

    fila_grupo = 8
    fila_encabezado = 9
    for col_ini, col_fin, texto in (grupos or []):
        if col_fin > col_ini:
            ws.merge_cells(start_row=fila_grupo, start_column=col_ini, end_row=fila_grupo, end_column=col_fin)
        celda = ws.cell(row=fila_grupo, column=col_ini, value=texto)
        celda.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=VERDE)
        celda.alignment = Alignment(horizontal="center", vertical="top")
        for j in range(col_ini, col_fin + 1):
            ws.cell(row=fila_grupo, column=j).border = Border(left=THIN_BLANCO, right=THIN_BLANCO)

    ws.row_dimensions[fila_encabezado].height = 30
    for j, texto in enumerate(encabezados, start=1):
        celda = ws.cell(row=fila_encabezado, column=j, value=texto)
        celda.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=BURDEOS)
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = BORDE_ENCABEZADO

    n_filas_agrupadoras = max(len(filas), 1)
    for i, (_, fila) in enumerate(pivote.iterrows()):
        r = fila_encabezado + 1 + i
        es_total = str(fila.iloc[0]) == "Total general"
        for j, col in enumerate(pivote.columns, start=1):
            valor = fila[col]
            celda = ws.cell(row=r, column=j, value=valor)
            celda.border = BORDE
            celda.alignment = Alignment(vertical="center", horizontal="center" if col in (filas or []) else None)
            if isinstance(valor, (int, float)) and col not in (filas or []):
                celda.number_format = FORMATO_MONEDA
            if es_total:
                celda.font = Font(name="Arial", size=11, bold=True)
                celda.fill = PatternFill("solid", fgColor=GRIS_TOTAL)
            else:
                celda.font = Font(name="Calibri", size=11)

    ws.freeze_panes = ws.cell(row=fila_encabezado + 2, column=n_filas_agrupadoras + 1)
    ws.print_title_rows = f"1:{fila_encabezado}"
    for j, texto in enumerate(encabezados, start=1):
        ancho = max(13, min(40, len(str(texto)) + 6))
        ws.column_dimensions[get_column_letter(j)].width = ancho

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def fecha_desde_nombre_archivo(nombre: str) -> str | None:
    m = re.search(r"(\d{1,2})[-_]([A-ZÁÉÍÓÚa-záéíóú]+)[-_](\d{4})", nombre)
    if not m:
        return None
    dia, mes_txt, anio = m.groups()
    mes_txt = mes_txt.lower()
    meses_largos = ["enero", "febrero", "marzo", "abril", "mayo", "junio", "julio",
                    "agosto", "septiembre", "octubre", "noviembre", "diciembre"]
    mes_encontrado = next((m for m in meses_largos if m.startswith(mes_txt[:3])), None)
    if not mes_encontrado:
        return None
    return f"{int(dia)} de {mes_encontrado} de {anio}"


# ---------------------------------------------------------------------------
# Interfaz Streamlit
# ---------------------------------------------------------------------------
def main():
    st.set_page_config(page_title="Tabla dinámica MAP / SICOP", layout="wide")
    st.title("Constructor de reportes — MAP / SICOP")
    st.caption(
        "Integra reportes MAP y SICOP en un solo lugar, arma cualquier reporte "
        "tipo tabla dinámica y descárgalo con el formato del Estado del Ejercicio."
    )

    with st.sidebar:
        st.header("1. Fuente de datos")
        fuente = st.radio("¿Qué vas a cargar?", ["MAP", "SICOP"], horizontal=True)
        archivo = st.file_uploader(f"Archivo crudo de {fuente} (.csv o .xlsx)", type=["csv", "xlsx", "xls"])
        st.divider()
        st.header("2. Periodo")
        hoy = date.today()
        mes_corte_idx = st.selectbox(
            "Corte de 'Al periodo' (acumulado enero → este mes)",
            options=list(range(12)),
            format_func=lambda i: NOMBRES_MES[i].capitalize(),
            index=min(hoy.month - 1, 11),
        )
        depurar_sicop = False
        if fuente == "SICOP":
            depurar_sicop = st.checkbox(
                "Excluir capítulo 1000, partida 39801 y CONTROL_OPERATIVO 60-69",
                value=False,
                help="Reglas confirmadas en tu procesador SICOP (nrbeca/nuevo). Revisa si aplican al reporte que quieres armar.",
            )

    if not archivo:
        st.info("Sube un archivo en la barra lateral para empezar a armar tu reporte.")
        return

    try:
        df = cargar_crudo(archivo.getvalue(), archivo.name, fuente)
    except Exception as e:
        st.error(f"No se pudo leer el archivo: {e}")
        return

    df = enriquecer_con_catalogos(df)
    df, columnas_familia = agregar_periodos_y_disponible(df, fuente, mes_corte_idx)
    if fuente == "SICOP" and depurar_sicop:
        df = aplicar_depuracion_sicop(df)

    st.success(f"Archivo cargado: {archivo.name} — {len(df):,} filas")

    columnas_no_valor = [c for c in df.columns if c not in columnas_familia]
    columnas_num_extra = [c for c in columnas_no_valor if pd.api.types.is_numeric_dtype(df[c])]
    columnas_cat = [c for c in columnas_no_valor if c not in columnas_num_extra]
    campos_sugeridos = [c for c in ["Unidad Responsable (nombre)", "Partida (nombre)", "Programa (nombre)", "Capítulo (nombre)"] if c in columnas_cat]

    fecha_archivo = fecha_desde_nombre_archivo(archivo.name)
    titulo_default = (f"Estado del Ejercicio al {fecha_archivo}" if fecha_archivo
                       else f"Estado del Ejercicio al {hoy.day} de {NOMBRES_MES[hoy.month-1]} de {hoy.year}")

    tab_estandar, tab_personalizado = st.tabs([" Reporte estándar (formato original)", " Reporte personalizado"])

    # -----------------------------------------------------------------
    # Reporte estándar: siempre sale con el formato de formato_estado_del_ejercicio.xlsx
    # -----------------------------------------------------------------
    with tab_estandar:
        st.caption("Este reporte sale automáticamente con el mismo formato que compartiste, sin importar si cargaste MAP o SICOP.")
        pivote_std, encabezados_std, grupos_std, filas_std = construir_reporte_plantilla(df, fuente, mes_corte_idx)
        if pivote_std.empty:
            st.warning("La base cargada no trae las columnas necesarias (Unidad Responsable / Partida) para armar el formato estándar.")
        else:
            st.dataframe(pivote_std, use_container_width=True, height=420)
            colA, colB = st.columns(2)
            with colA:
                linea1_std = st.text_input("Encabezado — línea 1", value="Unidad de Administración y Finanzas", key="linea1_std")
                titulo_std = st.text_input("Título del reporte", value=titulo_default, key="titulo_std")
            with colB:
                linea2_std = st.text_input("Encabezado — línea 2", value="Dirección General de Programación, Presupuesto y Finanzas", key="linea2_std")
                subtitulo_std = st.text_input("Subtítulo del reporte", value=f"Reporte {fuente}", key="subtitulo_std")

            excel_std = exportar_excel_oref(pivote_std, fuente, linea1_std, linea2_std, titulo_std, subtitulo_std,
                                             filas_std, encabezados_std, grupos_std)
            st.download_button(
                " Descargar Excel — formato original",
                data=excel_std,
                file_name=f"Estado_del_Ejercicio_{fuente}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # -----------------------------------------------------------------
    # Reporte personalizado: parte de los mismos campos, pero se puede
    # agregar o quitar cualquier cosa de la base para armar otro reporte.
    # -----------------------------------------------------------------
    with tab_personalizado:
        st.caption("Arranca con los mismos campos del reporte estándar y agrega o quita lo que necesites.")
        c1, c2, c3 = st.columns(3)
        with c1:
            filas = st.multiselect("Filas (agrupar por)", columnas_cat,
                                    default=[c for c in filas_std if c in columnas_cat] or campos_sugeridos[:1])
        with c2:
            columnas_pivote = st.multiselect("Columnas (opcional, para pivotear)", [c for c in columnas_cat if c not in filas])
        with c3:
            mostrar_mensuales = st.checkbox("Incluir columnas mensuales individuales en 'Valores'", value=False)
            opciones_valor = columnas_familia + (columnas_num_extra if mostrar_mensuales else [])
            prioridad = ["Original", "Reservas", "Modificado", "Comprometido", "Ejercido", "Disponible"]
            default_valores = [c for base in prioridad for c in columnas_familia if c.startswith(base + " (")]
            valores = st.multiselect("Valores (se suman)", opciones_valor, default=default_valores)

        with st.expander("Filtros — cualquier columna de la base", expanded=False):
            filtros = {}
            cols_filtro = st.multiselect("¿Qué columnas quieres filtrar?", columnas_cat, default=campos_sugeridos, key="cols_filtro_custom")
            for col in cols_filtro:
                opciones = sorted(df[col].dropna().astype(str).unique().tolist())
                filtros[col] = st.multiselect(f"Valores de «{col}»", opciones, key=f"filtro_custom_{col}")

        dff = df.copy()
        for col, seleccion in filtros.items():
            if seleccion:
                dff = dff[dff[col].astype(str).isin(seleccion)]

        pivote = construir_pivote(dff, filas, columnas_pivote, valores, {})

        st.subheader("Vista previa")
        if pivote.empty:
            st.warning("Elige al menos una columna en Filas y una en Valores para ver la tabla.")
        else:
            st.dataframe(pivote, use_container_width=True, height=420)

            colA, colB = st.columns(2)
            with colA:
                linea1 = st.text_input("Encabezado — línea 1", value="Unidad de Administración y Finanzas", key="linea1_custom")
                titulo = st.text_input("Título del reporte", value=titulo_default, key="titulo_custom")
            with colB:
                linea2 = st.text_input("Encabezado — línea 2", value="Dirección General de Programación, Presupuesto y Finanzas", key="linea2_custom")
                subtitulo = st.text_input("Subtítulo del reporte", value=f"Reporte {fuente} — datos seleccionados", key="subtitulo_custom")

            excel_bytes = exportar_excel_oref(pivote, fuente, linea1, linea2, titulo, subtitulo, filas)
            st.download_button(
                " Descargar Excel — reporte personalizado",
                data=excel_bytes,
                file_name=f"Reporte_{fuente}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )


if __name__ == "__main__":
    main()
